from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

TRACKER_SHEET = "Tracker"
LOG_SHEET = "Log"

TRACKER_COLUMNS = [
    "Company",
    "Contact Name",
    "Contact Info",
    "Source",
    "Problem Hypothesis",
    "Contact Method Used",
    "Date Contacted",
    "Response",
    "Actual Problem (confirmed)",
    "Proposed Solution",
    "Status",
    "Next Step",
    "Notes",
]

LOG_COLUMNS = [
    "Date & Time",
    "Company",
    "Contact Method",
    "Direction",
    "Outcome / What Happened",
]

VALID_STATUSES = [
    "New",
    "Contacted",
    "Discovery Call Done",
    "Proposal Sent",
    "Won",
    "Lost",
]


@dataclass
class TrackerRow:
    """One row of the Tracker sheet, as plain data."""

    company: str
    contact_name: str | None = None
    contact_info: str | None = None
    source: str | None = None
    problem_hypothesis: str | None = None
    contact_method_used: str | None = None
    date_contacted: str | None = None
    response: str | None = None
    actual_problem: str | None = None
    proposed_solution: str | None = None
    status: str = "New"
    next_step: str | None = None
    notes: str | None = None
    row_number: int | None = None  # 1-based Excel row, set when read from a file

    def as_dict(self) -> dict:
        return {
            "Company": self.company,
            "Contact Name": self.contact_name,
            "Contact Info": self.contact_info,
            "Source": self.source,
            "Problem Hypothesis": self.problem_hypothesis,
            "Contact Method Used": self.contact_method_used,
            "Date Contacted": self.date_contacted,
            "Response": self.response,
            "Actual Problem (confirmed)": self.actual_problem,
            "Proposed Solution": self.proposed_solution,
            "Status": self.status,
            "Next Step": self.next_step,
            "Notes": self.notes,
        }


@dataclass
class LogRow:
    """One row of the Log sheet - a single call/email/chat touchpoint."""

    company: str
    outcome: str
    contact_method: str = "Phone"
    direction: str = "Outbound"
    when: datetime = field(default_factory=datetime.now)


def ensure_workbook(path: str | Path) -> Path:
    """Create the workbook with correct headers if it doesn't already exist.

    If a file already exists at `path`, it is left untouched (including
    any extra columns or sheets the user has added) - this only fills in
    a blank starting point.
    """
    path = Path(path)
    if path.exists():
        return path

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    tracker_ws = wb.active
    tracker_ws.title = TRACKER_SHEET
    tracker_ws.append(TRACKER_COLUMNS)

    log_ws = wb.create_sheet(LOG_SHEET)
    log_ws.append(LOG_COLUMNS)

    wb.save(path)
    return path


def _header_index_map(ws: Worksheet, expected_columns: list[str]) -> dict[str, int]:
    """Map column name -> 1-based column index, reading the header row.

    Tolerant of the user's columns being in a different order than
    TRACKER_COLUMNS/LOG_COLUMNS, as long as the names match.
    """
    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    mapping: dict[str, int] = {}
    for cell in header_row:
        if cell.value:
            mapping[str(cell.value).strip()] = cell.column

    missing = [c for c in expected_columns if c not in mapping]
    if missing:
        raise ValueError(
            f"Tracker sheet '{ws.title}' is missing expected column(s): {missing}. "
            f"Found columns: {list(mapping)}"
        )
    return mapping


class TrackerSync:
    """Reads and writes a single Campaign_Tracker.xlsx file."""

    def __init__(self, path: str | Path):
        self.path = ensure_workbook(path)

    # ------------------------------------------------------------------ #
    # Reading
    # ------------------------------------------------------------------ #
    def read_rows(self) -> list[TrackerRow]:
        wb = load_workbook(self.path)
        ws = wb[TRACKER_SHEET]
        cols = _header_index_map(ws, TRACKER_COLUMNS)

        rows: list[TrackerRow] = []
        for row_cells in ws.iter_rows(min_row=2):
            row_num = row_cells[0].row

            def get(col_name: str):
                idx = cols[col_name] - 1
                return row_cells[idx].value

            company = get("Company")
            if not company:
                continue  # skip blank rows

            rows.append(
                TrackerRow(
                    company=str(company).strip(),
                    contact_name=get("Contact Name"),
                    contact_info=get("Contact Info"),
                    source=get("Source"),
                    problem_hypothesis=get("Problem Hypothesis"),
                    contact_method_used=get("Contact Method Used"),
                    date_contacted=get("Date Contacted"),
                    response=get("Response"),
                    actual_problem=get("Actual Problem (confirmed)"),
                    proposed_solution=get("Proposed Solution"),
                    status=get("Status") or "New",
                    next_step=get("Next Step"),
                    notes=get("Notes"),
                    row_number=row_num,
                )
            )
        return rows

    def find_row(self, company: str) -> TrackerRow | None:
        target = company.strip().casefold()
        for row in self.read_rows():
            if row.company.strip().casefold() == target:
                return row
        return None

    # ------------------------------------------------------------------ #
    # Writing
    # ------------------------------------------------------------------ #
    def upsert_row(self, row: TrackerRow) -> int:
        """Insert or update one Tracker row by Company name. Returns the row number."""
        wb = load_workbook(self.path)
        ws = wb[TRACKER_SHEET]
        cols = _header_index_map(ws, TRACKER_COLUMNS)

        existing_row_num = None
        target = row.company.strip().casefold()
        for r in ws.iter_rows(min_row=2):
            company_cell = r[cols["Company"] - 1]
            if company_cell.value and str(company_cell.value).strip().casefold() == target:
                existing_row_num = r[0].row
                break

        row_num = existing_row_num or (ws.max_row + 1 if ws.max_row >= 1 else 2)
        data = row.as_dict()
        for col_name, value in data.items():
            if value is None:
                continue
            ws.cell(row=row_num, column=cols[col_name], value=value)

        wb.save(self.path)
        return row_num

    def append_log_row(self, log_row: LogRow) -> int:
        wb = load_workbook(self.path)
        ws = wb[LOG_SHEET]
        cols = _header_index_map(ws, LOG_COLUMNS)

        row_num = ws.max_row + 1 if ws.max_row >= 1 else 2
        values = {
            "Date & Time": log_row.when.strftime("%Y-%m-%d %H:%M"),
            "Company": log_row.company,
            "Contact Method": log_row.contact_method,
            "Direction": log_row.direction,
            "Outcome / What Happened": log_row.outcome,
        }
        for col_name, value in values.items():
            ws.cell(row=row_num, column=cols[col_name], value=value)

        wb.save(self.path)
        return row_num

    def read_log_rows(self) -> list[dict]:
        wb = load_workbook(self.path)
        ws = wb[LOG_SHEET]
        cols = _header_index_map(ws, LOG_COLUMNS)

        rows = []
        for row_cells in ws.iter_rows(min_row=2):
            if not row_cells[0].value:
                continue
            rows.append({name: row_cells[idx - 1].value for name, idx in cols.items()})
        return rows